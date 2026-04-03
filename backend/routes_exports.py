"""
Module d'exports PDF et Excel pour Édu-Connect
Génère des rapports et exports de données pour tous les modules
"""
from fastapi import APIRouter, HTTPException, Depends, status
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime, timezone
from io import BytesIO
import os

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.pdfgen import canvas

from auth import get_current_user
from dotenv import load_dotenv
from pathlib import Path

# Router
router = APIRouter(prefix="/api/exports", tags=["Exports"])

# MongoDB connection
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]


# ============================================
# HELPERS - Excel
# ============================================

def create_excel_workbook(title: str):
    """Crée un classeur Excel avec un style professionnel"""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limite à 31 caractères
    return wb, ws


def style_header(ws, headers: list, row: int = 1):
    """Applique un style professionnel aux en-têtes"""
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = 20


def add_metadata(ws, title: str):
    """Ajoute des métadonnées en haut du fichier Excel"""
    ws.insert_rows(1, 3)
    
    # Titre
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, color="1F4788")
    
    # Date d'export
    ws['A2'] = f"Exporté le : {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)
    
    # Ligne vide
    ws.row_dimensions[3].height = 5


# ============================================
# EXPORTS - Statistiques Dashboard
# ============================================

@router.get(
    "/dashboard/stats",
    summary="Exporter les statistiques du dashboard",
    description="Génère un fichier Excel avec toutes les statistiques du dashboard"
)
async def export_dashboard_stats(
    current_user: dict = Depends(get_current_user)
):
    """Export Excel des statistiques du dashboard"""
    
    # Récupérer les statistiques
    total_etablissements = await db.etablissements.count_documents({})
    total_enseignants = await db.enseignants.count_documents({})
    total_eleves = await db.eleves.count_documents({})
    total_classes = await db.classes.count_documents({})
    
    # Créer le classeur Excel
    wb, ws = create_excel_workbook("Statistiques")
    add_metadata(ws, "📊 Statistiques Générales - Édu-Connect")
    
    # En-têtes
    headers = ["Indicateur", "Valeur", "Détails"]
    style_header(ws, headers, row=4)
    
    # Données
    data = [
        ["🏫 Établissements", total_etablissements, f"Total des écoles enregistrées"],
        ["👨‍🏫 Enseignants", total_enseignants, f"Personnel enseignant actif"],
        ["👨‍🎓 Élèves", total_eleves, f"Élèves inscrits"],
        ["🎓 Classes", total_classes, f"Classes ouvertes"],
    ]
    
    for row_num, row_data in enumerate(data, start=5):
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal='center' if col_num == 2 else 'left')
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    
    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"educonnect_stats_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================
# EXPORTS - Enseignants
# ============================================

@router.get(
    "/enseignants",
    summary="Exporter la liste des enseignants",
    description="Génère un fichier Excel avec la liste complète des enseignants"
)
async def export_enseignants(
    current_user: dict = Depends(get_current_user)
):
    """Export Excel de la liste des enseignants"""
    
    # Récupérer tous les enseignants
    enseignants = await db.enseignants.find({}, {"_id": 0}).to_list(10000)
    
    if not enseignants:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Aucun enseignant à exporter"
        )
    
    # Créer le classeur Excel
    wb, ws = create_excel_workbook("Enseignants")
    add_metadata(ws, "👨‍🏫 Liste des Enseignants - Édu-Connect")
    
    # En-têtes
    headers = [
        "Matricule", "Nom", "Prénom", "Sexe", "Téléphone", 
        "Email", "Grade", "Discipline", "Établissement"
    ]
    style_header(ws, headers, row=4)
    
    # Données
    for row_num, ens in enumerate(enseignants, start=5):
        ws.cell(row=row_num, column=1).value = ens.get("matricule", "N/A")
        ws.cell(row=row_num, column=2).value = ens.get("nom", "")
        ws.cell(row=row_num, column=3).value = ens.get("prenom", "")
        ws.cell(row=row_num, column=4).value = ens.get("sexe", "")
        ws.cell(row=row_num, column=5).value = ens.get("telephone", "")
        ws.cell(row=row_num, column=6).value = ens.get("email", "")
        ws.cell(row=row_num, column=7).value = ens.get("grade", "")
        ws.cell(row=row_num, column=8).value = ens.get("discipline_principale", "")
        ws.cell(row=row_num, column=9).value = ens.get("etablissement_actuel", "")
    
    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"educonnect_enseignants_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================
# EXPORTS - Élèves
# ============================================

@router.get(
    "/eleves",
    summary="Exporter la liste des élèves",
    description="Génère un fichier Excel avec la liste complète des élèves"
)
async def export_eleves(
    current_user: dict = Depends(get_current_user)
):
    """Export Excel de la liste des élèves"""
    
    # Récupérer tous les élèves
    eleves = await db.eleves.find({}, {"_id": 0}).to_list(10000)
    
    if not eleves:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Aucun élève à exporter"
        )
    
    # Créer le classeur Excel
    wb, ws = create_excel_workbook("Élèves")
    add_metadata(ws, "👨‍🎓 Liste des Élèves - Édu-Connect")
    
    # En-têtes
    headers = [
        "INE", "Nom", "Prénom", "Sexe", "Date Naissance",
        "Niveau", "Classe", "Établissement", "Statut"
    ]
    style_header(ws, headers, row=4)
    
    # Données
    for row_num, eleve in enumerate(eleves, start=5):
        ws.cell(row=row_num, column=1).value = eleve.get("ine", "N/A")
        ws.cell(row=row_num, column=2).value = eleve.get("nom", "")
        ws.cell(row=row_num, column=3).value = eleve.get("prenom", "")
        ws.cell(row=row_num, column=4).value = eleve.get("sexe", "")
        
        date_naissance = eleve.get("date_naissance", "")
        if isinstance(date_naissance, str):
            ws.cell(row=row_num, column=5).value = date_naissance
        
        ws.cell(row=row_num, column=6).value = eleve.get("niveau_scolaire", "")
        ws.cell(row=row_num, column=7).value = eleve.get("classe_id", "")
        ws.cell(row=row_num, column=8).value = eleve.get("etablissement_id", "")
        ws.cell(row=row_num, column=9).value = "Actif" if eleve.get("is_active", True) else "Inactif"
    
    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"educonnect_eleves_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================
# EXPORTS - Établissements
# ============================================

@router.get(
    "/etablissements",
    summary="Exporter la liste des établissements",
    description="Génère un fichier Excel avec la liste complète des établissements"
)
async def export_etablissements(
    current_user: dict = Depends(get_current_user)
):
    """Export Excel de la liste des établissements"""
    
    # Récupérer tous les établissements
    etablissements = await db.etablissements.find({}, {"_id": 0}).to_list(10000)
    
    if not etablissements:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Aucun établissement à exporter"
        )
    
    # Créer le classeur Excel
    wb, ws = create_excel_workbook("Établissements")
    add_metadata(ws, "🏫 Liste des Établissements - Édu-Connect")
    
    # En-têtes
    headers = [
        "Code", "Nom", "Type", "Province", "Ville",
        "Adresse", "Téléphone", "Email", "Statut"
    ]
    style_header(ws, headers, row=4)
    
    # Données
    for row_num, etab in enumerate(etablissements, start=5):
        ws.cell(row=row_num, column=1).value = etab.get("code", "N/A")
        ws.cell(row=row_num, column=2).value = etab.get("nom", "")
        ws.cell(row=row_num, column=3).value = etab.get("type", "")
        ws.cell(row=row_num, column=4).value = etab.get("province", "")
        ws.cell(row=row_num, column=5).value = etab.get("ville", "")
        ws.cell(row=row_num, column=6).value = etab.get("adresse", "")
        ws.cell(row=row_num, column=7).value = etab.get("telephone", "")
        ws.cell(row=row_num, column=8).value = etab.get("email", "")
        ws.cell(row=row_num, column=9).value = "Actif" if etab.get("is_active", True) else "Inactif"
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    
    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"educonnect_etablissements_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================
# EXPORTS - Documents GED
# ============================================

@router.get(
    "/documents",
    summary="Exporter la liste des documents",
    description="Génère un fichier Excel avec la liste des documents GED"
)
async def export_documents(
    current_user: dict = Depends(get_current_user)
):
    """Export Excel de la liste des documents GED"""
    
    # Récupérer tous les documents
    documents = await db.documents.find({}, {"_id": 0}).to_list(10000)
    
    if not documents:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Aucun document à exporter"
        )
    
    # Créer le classeur Excel
    wb, ws = create_excel_workbook("Documents")
    add_metadata(ws, "📄 Documents GED - Édu-Connect")
    
    # En-têtes
    headers = [
        "Titre", "Type", "Statut", "Priorité", "Service Créateur",
        "Date Création", "Destinataires", "Objet"
    ]
    style_header(ws, headers, row=4)
    
    # Données
    for row_num, doc in enumerate(documents, start=5):
        ws.cell(row=row_num, column=1).value = doc.get("titre", "")
        ws.cell(row=row_num, column=2).value = doc.get("type_document", "")
        ws.cell(row=row_num, column=3).value = doc.get("statut", "")
        ws.cell(row=row_num, column=4).value = doc.get("priorite", "")
        ws.cell(row=row_num, column=5).value = doc.get("service_createur_nom", "")
        
        date_creation = doc.get("date_creation", "")
        if isinstance(date_creation, str):
            ws.cell(row=row_num, column=6).value = date_creation[:10]
        
        destinataires = doc.get("destinataires", [])
        ws.cell(row=row_num, column=7).value = ", ".join(destinataires[:3]) if destinataires else "N/A"
        ws.cell(row=row_num, column=8).value = doc.get("objet", "")[:50]
    
    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['H'].width = 40
    
    # Sauvegarder
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"educonnect_documents_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
